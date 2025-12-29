from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func
from typing import Optional
from datetime import datetime, date, timedelta
import csv
import io
import logging

from app.core.database import get_db
from app.api.v1.auth import get_current_user
from app.models.user import User
from app.models.ticket import Ticket
from app.models.report import ReportSnapshot, PeriodType
from app.schemas.report import (
    ReportSummaryResponse,
    SnapshotResponse,
    SnapshotDetailResponse,
    SnapshotListResponse,
    SnapshotCreate,
    SnapshotGenerateResponse,
    SchedulerStatusResponse,
    PeriodType as SchemaPeriodType,
    ReportStatsResponse,
    ReportTrendsResponse,
    ReportDistributionResponse,
    DistributionItem,
    ReportSnapshotsResponse,
    SnapshotTableRow,
)
from app.services.report_service import ReportService
from app.jobs.report_batch import get_scheduler_status
from app.models.user import UserRole

logger = logging.getLogger(__name__)

router = APIRouter()


@router.get("/summary", response_model=ReportSummaryResponse)
async def get_report_summary(
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get report summary for a date range."""
    # Build query
    query = select(Ticket).where(Ticket.tenant_id == current_user.tenant_id)

    if from_date:
        query = query.where(Ticket.created_at >= datetime.combine(from_date, datetime.min.time()))
    if to_date:
        query = query.where(Ticket.created_at <= datetime.combine(to_date, datetime.max.time()))

    result = await db.execute(query)
    tickets = result.scalars().all()

    # Calculate metrics
    total_tickets = len(tickets)
    by_status = {}
    by_priority = {}
    by_category = {}
    sla_breached = 0

    for ticket in tickets:
        # Count by status
        status_key = ticket.current_status.value
        by_status[status_key] = by_status.get(status_key, 0) + 1

        # Count by priority
        priority_key = ticket.priority.value
        by_priority[priority_key] = by_priority.get(priority_key, 0) + 1

        # Count by category
        category_key = ticket.category.value
        by_category[category_key] = by_category.get(category_key, 0) + 1

        # Count SLA breaches
        if ticket.sla_breached:
            sla_breached += 1

    # Calculate average resolution time (for closed tickets)
    closed_tickets = [t for t in tickets if t.closed_at and t.opened_at]
    avg_resolution_hours = 0
    if closed_tickets:
        total_hours = sum(
            (t.closed_at - t.opened_at).total_seconds() / 3600
            for t in closed_tickets
        )
        avg_resolution_hours = round(total_hours / len(closed_tickets), 2)

    return {
        "total_tickets": total_tickets,
        "by_status": by_status,
        "by_priority": by_priority,
        "by_category": by_category,
        "avg_resolution_time_hours": avg_resolution_hours,
        "sla_breached": sla_breached,
        "sla_compliance_rate": round(1 - (sla_breached / total_tickets), 3) if total_tickets > 0 else 1.0
    }


@router.get("/stats", response_model=ReportStatsResponse)
async def get_report_stats(
    start_date: date = Query(..., description="Start date for the report period"),
    end_date: date = Query(..., description="End date for the report period"),
    period: str = Query("daily", description="Period type: daily, weekly, monthly"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get report statistics with trend comparison."""
    # Calculate period length for comparison
    period_length = (end_date - start_date).days
    prev_start = start_date - timedelta(days=period_length)
    prev_end = start_date - timedelta(days=1)

    # Current period query
    current_query = select(Ticket).where(
        and_(
            Ticket.tenant_id == current_user.tenant_id,
            Ticket.created_at >= datetime.combine(start_date, datetime.min.time()),
            Ticket.created_at <= datetime.combine(end_date, datetime.max.time())
        )
    )
    current_result = await db.execute(current_query)
    current_tickets = current_result.scalars().all()

    # Previous period query
    prev_query = select(Ticket).where(
        and_(
            Ticket.tenant_id == current_user.tenant_id,
            Ticket.created_at >= datetime.combine(prev_start, datetime.min.time()),
            Ticket.created_at <= datetime.combine(prev_end, datetime.max.time())
        )
    )
    prev_result = await db.execute(prev_query)
    prev_tickets = prev_result.scalars().all()

    # Calculate current metrics
    total_tickets = len(current_tickets)
    resolved_count = sum(1 for t in current_tickets if t.current_status.value in ['resolved', 'closed'])
    sla_breached = sum(1 for t in current_tickets if t.sla_breached)
    sla_breach_rate = round((sla_breached / total_tickets * 100), 2) if total_tickets > 0 else 0

    closed_tickets = [t for t in current_tickets if t.closed_at and t.opened_at]
    avg_resolution_time = 0
    if closed_tickets:
        total_hours = sum((t.closed_at - t.opened_at).total_seconds() / 3600 for t in closed_tickets)
        avg_resolution_time = round(total_hours / len(closed_tickets), 2)

    # Calculate previous metrics
    prev_total = len(prev_tickets)
    prev_resolved = sum(1 for t in prev_tickets if t.current_status.value in ['resolved', 'closed'])
    prev_sla_breached = sum(1 for t in prev_tickets if t.sla_breached)
    prev_sla_rate = (prev_sla_breached / prev_total * 100) if prev_total > 0 else 0

    prev_closed = [t for t in prev_tickets if t.closed_at and t.opened_at]
    prev_avg_time = 0
    if prev_closed:
        prev_total_hours = sum((t.closed_at - t.opened_at).total_seconds() / 3600 for t in prev_closed)
        prev_avg_time = prev_total_hours / len(prev_closed)

    # Calculate trends
    total_trend = round(((total_tickets - prev_total) / prev_total * 100), 2) if prev_total > 0 else 0
    resolved_trend = round(((resolved_count - prev_resolved) / prev_resolved * 100), 2) if prev_resolved > 0 else 0
    sla_trend = round((sla_breach_rate - prev_sla_rate), 2)
    time_trend = round(((avg_resolution_time - prev_avg_time) / prev_avg_time * 100), 2) if prev_avg_time > 0 else 0

    return ReportStatsResponse(
        total_tickets=total_tickets,
        total_trend=total_trend,
        resolved_count=resolved_count,
        resolved_trend=resolved_trend,
        sla_breach_rate=sla_breach_rate,
        sla_trend=sla_trend,
        avg_resolution_time=avg_resolution_time,
        time_trend=time_trend
    )


@router.get("/trends", response_model=ReportTrendsResponse)
async def get_report_trends(
    start_date: date = Query(..., description="Start date for the report period"),
    end_date: date = Query(..., description="End date for the report period"),
    period: str = Query("daily", description="Period type: daily, weekly, monthly"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get ticket volume trends over time."""
    # Query tickets in the period
    query = select(Ticket).where(
        and_(
            Ticket.tenant_id == current_user.tenant_id,
            Ticket.created_at >= datetime.combine(start_date, datetime.min.time()),
            Ticket.created_at <= datetime.combine(end_date, datetime.max.time())
        )
    ).order_by(Ticket.created_at)
    
    result = await db.execute(query)
    tickets = result.scalars().all()

    # Group tickets by date
    tickets_by_date = {}
    for ticket in tickets:
        ticket_date = ticket.created_at.date()
        tickets_by_date[ticket_date] = tickets_by_date.get(ticket_date, 0) + 1

    # Generate date range
    labels = []
    values = []
    current = start_date
    while current <= end_date:
        labels.append(current.strftime('%Y-%m-%d'))
        values.append(tickets_by_date.get(current, 0))
        current += timedelta(days=1)

    return ReportTrendsResponse(labels=labels, values=values)


@router.get("/distribution", response_model=ReportDistributionResponse)
async def get_report_distribution(
    start_date: date = Query(..., description="Start date for the report period"),
    end_date: date = Query(..., description="End date for the report period"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get ticket distribution by category and priority."""
    # Query tickets in the period
    query = select(Ticket).where(
        and_(
            Ticket.tenant_id == current_user.tenant_id,
            Ticket.created_at >= datetime.combine(start_date, datetime.min.time()),
            Ticket.created_at <= datetime.combine(end_date, datetime.max.time())
        )
    )
    
    result = await db.execute(query)
    tickets = result.scalars().all()

    total = len(tickets)
    if total == 0:
        return ReportDistributionResponse(by_category=[], by_priority=[])

    # Count by category
    category_counts = {}
    for ticket in tickets:
        cat = ticket.category.value
        category_counts[cat] = category_counts.get(cat, 0) + 1

    by_category = [
        DistributionItem(
            name=cat,
            count=count,
            percentage=round((count / total) * 100, 2)
        )
        for cat, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True)
    ]

    # Count by priority
    priority_counts = {}
    for ticket in tickets:
        pri = ticket.priority.value
        priority_counts[pri] = priority_counts.get(pri, 0) + 1

    by_priority = [
        DistributionItem(
            name=pri,
            count=count,
            percentage=round((count / total) * 100, 2)
        )
        for pri, count in sorted(priority_counts.items(), key=lambda x: x[1], reverse=True)
    ]

    return ReportDistributionResponse(by_category=by_category, by_priority=by_priority)


@router.get("/snapshots/recent", response_model=ReportSnapshotsResponse)
async def get_recent_snapshots(
    limit: int = Query(10, ge=1, le=50, description="Number of recent snapshots to return"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get recent report snapshots for the table view."""
    report_service = ReportService(db)
    snapshots, _ = await report_service.list_snapshots(
        tenant_id=current_user.tenant_id,
        period_type=None,
        from_date=None,
        to_date=None,
        skip=0,
        limit=limit
    )

    # Convert to table rows
    rows = []
    for snapshot in snapshots:
        metrics = snapshot.metrics or {}
        total = metrics.get('total_created', 0)
        resolved = metrics.get('total_resolved', 0)
        breached = metrics.get('sla_breached_count', 0)
        avg_hours = metrics.get('avg_resolution_time_hours', 0)
        
        rows.append(SnapshotTableRow(
            id=snapshot.id,
            date=snapshot.period_start.strftime('%Y-%m-%d'),
            total=total,
            resolved=resolved,
            breached=breached,
            avgTime=int(avg_hours * 60)  # Convert hours to minutes
        ))

    return ReportSnapshotsResponse(snapshots=rows)


@router.get("/export/csv")
async def export_tickets_csv(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Export tickets to CSV."""
    # Build query
    query = select(Ticket).where(Ticket.tenant_id == current_user.tenant_id)

    if start_date:
        query = query.where(Ticket.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.where(Ticket.created_at <= datetime.combine(end_date, datetime.max.time()))

    query = query.order_by(Ticket.created_at.desc())

    result = await db.execute(query)
    tickets = result.scalars().all()

    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow([
        'Ticket Number', 'Title', 'Status', 'Priority', 'Category',
        'Created At', 'Closed At', 'SLA Breached'
    ])

    # Write data
    for ticket in tickets:
        writer.writerow([
            ticket.ticket_number,
            ticket.title,
            ticket.current_status.value,
            ticket.priority.value,
            ticket.category.value,
            ticket.created_at.isoformat(),
            ticket.closed_at.isoformat() if ticket.closed_at else '',
            'Yes' if ticket.sla_breached else 'No'
        ])

    # Return CSV file
    output.seek(0)
    filename = f"tickets_export_{datetime.utcnow().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/xlsx")
async def export_tickets_xlsx(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Export tickets to Excel (XLSX)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel export requires openpyxl package"
        )

    # Build query
    query = select(Ticket).where(Ticket.tenant_id == current_user.tenant_id)

    if start_date:
        query = query.where(Ticket.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.where(Ticket.created_at <= datetime.combine(end_date, datetime.max.time()))

    query = query.order_by(Ticket.created_at.desc())

    result = await db.execute(query)
    tickets = result.scalars().all()

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # Header style
    header_fill = PatternFill(start_color="EB5D19", end_color="EB5D19", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Write header
    headers = ['Ticket Number', 'Title', 'Status', 'Priority', 'Category', 
               'Created At', 'Closed At', 'SLA Breached']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    for row_num, ticket in enumerate(tickets, 2):
        ws.cell(row=row_num, column=1, value=ticket.ticket_number)
        ws.cell(row=row_num, column=2, value=ticket.title)
        ws.cell(row=row_num, column=3, value=ticket.current_status.value)
        ws.cell(row=row_num, column=4, value=ticket.priority.value)
        ws.cell(row=row_num, column=5, value=ticket.category.value)
        ws.cell(row=row_num, column=6, value=ticket.created_at.isoformat())
        ws.cell(row=row_num, column=7, value=ticket.closed_at.isoformat() if ticket.closed_at else '')
        ws.cell(row=row_num, column=8, value='Yes' if ticket.sla_breached else 'No')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"tickets_export_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export")
async def export_tickets_csv_legacy(
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Legacy CSV export endpoint (deprecated, use /export/csv instead)."""
    return await export_tickets_csv(from_date, to_date, current_user, db)


# ============================================================================
# Report Snapshot Endpoints
# ============================================================================

@router.get("/snapshots", response_model=SnapshotListResponse)
async def list_snapshots(
    period_type: Optional[SchemaPeriodType] = Query(None, description="Filter by period type"),
    from_date: Optional[date] = Query(None, description="Filter snapshots starting from this date"),
    to_date: Optional[date] = Query(None, description="Filter snapshots ending before this date"),
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(50, ge=1, le=100, description="Maximum number of records to return"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    List report snapshots with optional filters.

    Returns paginated list of report snapshots for the current tenant.
    Can filter by period type (day, week, month) and date range.
    """
    # Convert schema enum to model enum if provided
    model_period_type = None
    if period_type:
        model_period_type = PeriodType(period_type.value)

    report_service = ReportService(db)
    snapshots, total = await report_service.list_snapshots(
        tenant_id=current_user.tenant_id,
        period_type=model_period_type,
        from_date=from_date,
        to_date=to_date,
        skip=skip,
        limit=limit
    )

    return SnapshotListResponse(
        items=[SnapshotResponse.model_validate(s) for s in snapshots],
        total=total,
        skip=skip,
        limit=limit
    )


@router.get("/snapshots/{snapshot_id}", response_model=SnapshotDetailResponse)
async def get_snapshot(
    snapshot_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Get a specific report snapshot by ID.

    Returns detailed snapshot information including parsed metrics
    with breakdown by status, priority, category, and top sites.
    """
    report_service = ReportService(db)
    snapshot = await report_service.get_snapshot_by_id(
        snapshot_id=snapshot_id,
        tenant_id=current_user.tenant_id
    )

    if not snapshot:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Snapshot with id {snapshot_id} not found"
        )

    return SnapshotDetailResponse.from_snapshot(snapshot)


@router.post("/snapshots/generate", response_model=SnapshotGenerateResponse)
async def generate_snapshot(
    request: SnapshotCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Manually generate a report snapshot.

    Generates a snapshot for the specified period type and date.
    If target_date is not provided, defaults to the previous period:
    - day: yesterday
    - week: previous week (Monday to Sunday)
    - month: previous month

    Requires admin or manager role.
    """
    # Check user permissions (admin or manager)
    if current_user.role not in [UserRole.ADMIN, UserRole.TENANT_ADMIN, UserRole.AS_MANAGER]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admin or manager users can manually generate snapshots"
        )

    report_service = ReportService(db)
    target_date = request.target_date

    try:
        if request.period_type == SchemaPeriodType.DAY:
            if target_date is None:
                target_date = date.today() - timedelta(days=1)
            snapshot = await report_service.generate_daily_snapshot(
                tenant_id=current_user.tenant_id,
                target_date=target_date
            )
            message = f"Daily snapshot generated for {target_date}"

        elif request.period_type == SchemaPeriodType.WEEK:
            if target_date is None:
                # Default to previous week's Monday
                today = date.today()
                days_since_monday = today.weekday()
                target_date = today - timedelta(days=days_since_monday + 7)
            snapshot = await report_service.generate_weekly_snapshot(
                tenant_id=current_user.tenant_id,
                week_start=target_date
            )
            message = f"Weekly snapshot generated for week starting {snapshot.period_start}"

        elif request.period_type == SchemaPeriodType.MONTH:
            if target_date is None:
                # Default to previous month
                today = date.today()
                first_of_current_month = date(today.year, today.month, 1)
                last_of_previous_month = first_of_current_month - timedelta(days=1)
                target_date = last_of_previous_month
            snapshot = await report_service.generate_monthly_snapshot(
                tenant_id=current_user.tenant_id,
                year=target_date.year,
                month=target_date.month
            )
            message = f"Monthly snapshot generated for {target_date.year}-{target_date.month:02d}"

        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid period type: {request.period_type}"
            )

        logger.info(
            f"User {current_user.id} manually generated {request.period_type} snapshot "
            f"for tenant {current_user.tenant_id}"
        )

        return SnapshotGenerateResponse(
            snapshot=SnapshotResponse.model_validate(snapshot),
            message=message
        )

    except Exception as e:
        logger.error(f"Failed to generate snapshot: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate snapshot: {str(e)}"
        )


@router.delete("/snapshots/{snapshot_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_snapshot(
    snapshot_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Delete a report snapshot.

    Requires admin role.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admin users can delete snapshots"
        )

    report_service = ReportService(db)
    deleted = await report_service.delete_snapshot(
        snapshot_id=snapshot_id,
        tenant_id=current_user.tenant_id
    )

    if not deleted:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Snapshot with id {snapshot_id} not found"
        )

    logger.info(
        f"User {current_user.id} deleted snapshot {snapshot_id} "
        f"for tenant {current_user.tenant_id}"
    )


@router.get("/scheduler/status", response_model=SchedulerStatusResponse)
async def get_report_scheduler_status(
    current_user: User = Depends(get_current_user)
):
    """
    Get the status of the report batch scheduler.

    Returns information about scheduled jobs including their next run time.
    Requires admin role.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.TENANT_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admin users can view scheduler status"
        )

    status_info = get_scheduler_status()
    return SchedulerStatusResponse(**status_info)
